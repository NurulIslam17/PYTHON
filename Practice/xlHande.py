
import openpyxl as xl
wb=xl.load_workbook('transaction.xlsx')
sheet=wb['Sheet1']
#cell=sheet['a1']# co ordinate of the sheet

cell=sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)

for row in range(1,sheet.max_row+1):
    print(row)


for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    print(cell.value)


# cell updated ..with new file
for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    corrected_value= cell.value * float(0.9)
    corrected_value_cell=sheet.cell(row,4)
    corrected_value_cell.value=corrected_value

wb.save('transaction2')

